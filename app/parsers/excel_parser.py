"""
Excel Parser - ИСПРАВЛЕНО
Теперь использует универсальный нормализатор для всех форматов
"""
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import re
import unicodedata

from app.utils.position_normalizer import normalize_positions
from app.core.registry import registry

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parse construction estimates from Excel files"""
    
    def parse(self, file_path: Path) -> Dict[str, Any]:
        """
        Parse Excel file and extract positions
        
        Args:
            file_path: Path to Excel file (.xlsx, .xls)
            
        Returns:
            {
                "document_info": {...},
                "positions": [...]
            }
        """
        logger.info(f"📊 Parsing Excel: {file_path.name}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            logger.info(f"Excel has {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
            
            all_positions = []
            sheet_summaries: List[Dict[str, Any]] = []
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                logger.info(f"Processing sheet: {sheet_name}")
                
                sheet = workbook[sheet_name]
                sheet_positions, sheet_diag = self._parse_sheet(sheet, sheet_name)

                raw_count = len(sheet_positions)

                if sheet_positions:
                    logger.info(
                        f"Extracted {raw_count} positions from sheet '{sheet_name}'"
                    )
                    all_positions.extend(sheet_positions)
                else:
                    logger.debug(f"No positions found in sheet '{sheet_name}'")

                sheet_summary = {
                    "sheet_name": sheet_name,
                    "raw_positions": raw_count,
                }
                sheet_summary.update(sheet_diag)

                sheet_summaries.append(sheet_summary)

            logger.info(f"Total raw positions from all sheets: {len(all_positions)}")

            # Normalize all positions and capture statistics
            normalized_positions, normalization_stats = normalize_positions(
                all_positions,
                return_stats=True
            )

            logger.info(
                f"✅ Excel parsed: {normalization_stats['normalized_total']} valid positions "
                f"from {len(workbook.sheetnames)} sheet(s)"
            )

            header_detection = [
                {
                    "sheet_name": summary.get("sheet_name"),
                    "header_found": summary.get("header_found", False),
                    "header_row": summary.get("header_row"),
                }
                for summary in sheet_summaries
            ]

            return {
                "document_info": {
                    "filename": file_path.name,
                    "format": "excel",
                    "sheets": workbook.sheetnames,
                    "total_sheets": len(workbook.sheetnames)
                },
                "positions": normalized_positions,
                "diagnostics": {
                    "raw_total": normalization_stats["raw_total"],
                    "normalized_total": normalization_stats["normalized_total"],
                    "skipped_total": normalization_stats["skipped_total"],
                    "sheet_summaries": sheet_summaries,
                    "header_detection": header_detection,
                }
            }

        except Exception as e:
            logger.error(f"❌ Excel parsing failed: {str(e)}", exc_info=True)
            return {
                "document_info": {
                    "filename": file_path.name,
                    "format": "excel",
                    "error": str(e)
                },
                "positions": [],
                "diagnostics": {
                    "raw_total": 0,
                    "normalized_total": 0,
                    "skipped_total": 0,
                    "sheet_summaries": [],
                    "header_detection": []
                }
            }

    def _parse_sheet(
        self, sheet: Worksheet, sheet_name: str
    ) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        """
        Parse a single Excel sheet
        
        Args:
            sheet: openpyxl Worksheet object
            sheet_name: Name of the sheet
            
        Returns:
            Tuple of raw position dicts (not normalized yet) and diagnostics
        """
        if not sheet or sheet.max_row < 1:
            logger.debug(f"Sheet '{sheet_name}' is empty or too small")
            diagnostics = {
                "header_found": False,
                "header_row": None,
                "header_values": [],
                "matched_keywords": [],
                "searched_rows": sheet.max_row if sheet else 0,
                "reason": "Sheet is empty or too small",
            }
            return [], diagnostics
        
        positions: List[Dict[str, Any]] = []
        skipped_entries: List[Dict[str, Any]] = []
        header_blocks: List[Dict[str, Any]] = []

        positions_skipped = 0
        last_data_row: Optional[int] = None

        active_header_info: List[Dict[str, Any]] = []
        active_header_keys: List[str] = []
        field_to_column: Dict[str, int] = {}

        first_header: Optional[Dict[str, Any]] = None
        best_candidate: Optional[Dict[str, Any]] = None
        sample_rows: List[List[str]] = []

        required_fields = {"description", "quantity"}

        def record_skip(row_index: int, reason: str, **extra: Any) -> None:
            nonlocal positions_skipped
            positions_skipped += 1
            entry: Dict[str, Any] = {"row": row_index, "reason": reason}
            for key, value in extra.items():
                if value is not None:
                    entry[key] = value
            skipped_entries.append(entry)

            extra_info = ", ".join(f"{key}={value}" for key, value in extra.items() if value is not None)
            logger.debug(
                "Skipping row %s in sheet '%s': %s%s",
                row_index,
                sheet_name,
                reason,
                f" ({extra_info})" if extra_info else "",
            )

        for row_idx in range(1, sheet.max_row + 1):
            row = sheet[row_idx]

            raw_values: List[str] = []
            normalized_values: List[str] = []
            has_values = False

            for cell in row:
                raw_value = ""
                if cell.value is not None:
                    if isinstance(cell.value, str):
                        raw_value = cell.value.strip()
                    else:
                        raw_value = str(cell.value).strip()

                if raw_value:
                    has_values = True

                raw_values.append(raw_value)
                normalized_values.append(ExcelParser._normalize_header_value(cell.value))

            if len(sample_rows) < 5:
                sample_rows.append(raw_values)

            candidate = self._detect_header_candidate(normalized_values)
            if candidate and (
                not best_candidate or candidate["match_count"] > best_candidate["match_count"]
            ):
                best_candidate = {
                    "row_index": row_idx,
                    "match_count": candidate["match_count"],
                    "matched_fields": candidate["matched_fields"],
                    "raw_headers": raw_values,
                }

            if candidate and candidate["is_header"]:
                active_header_info = []
                active_header_keys = []
                field_to_column = {}

                for col_idx, raw_header in enumerate(raw_values):
                    normalized_header = normalized_values[col_idx]
                    header_key = normalized_header.replace(" ", "_") if normalized_header else f"col_{col_idx}"
                    if header_key in active_header_keys:
                        header_key = f"{header_key}_{col_idx}"

                    canonical_field = self._map_header_to_field(normalized_header)

                    active_header_info.append({
                        "index": col_idx,
                        "raw": raw_header,
                        "normalized": normalized_header,
                        "field": canonical_field,
                        "header_key": header_key,
                    })
                    active_header_keys.append(header_key)

                    if canonical_field and canonical_field not in field_to_column:
                        field_to_column[canonical_field] = col_idx

                header_block = {
                    "row": row_idx,
                    "raw_headers": raw_values,
                    "normalized_headers": normalized_values,
                    "canonical_fields": [info["field"] for info in active_header_info],
                    "field_mapping": field_to_column.copy(),
                    "matched_fields": candidate["matched_fields"],
                }
                header_blocks.append(header_block)

                logger.info(
                    "🔎 Header detected in sheet '%s' at row %s: matched=%s, mapping=%s",
                    sheet_name,
                    row_idx,
                    candidate["matched_fields"],
                    field_to_column,
                )

                if first_header is None:
                    first_header = {
                        "row": row_idx,
                        "raw_headers": raw_values,
                        "matched_fields": candidate["matched_fields"],
                    }

                # Header rows should not be processed as data
                continue

            if not has_values:
                record_skip(row_idx, "empty_row")
                continue

            if not active_header_info:
                preview = " ".join(value for value in raw_values if value)[:120]
                record_skip(row_idx, "no_active_header", preview=preview)
                continue

            row_text = " ".join(value for value in raw_values if value)
            keyword_reason = self._detect_service_keyword(row_text.lower())
            if keyword_reason:
                record_skip(row_idx, "service_keyword", keyword=keyword_reason)
                continue

            position: Dict[str, Any] = {}
            canonical_row: Dict[str, Any] = {}

            for col_idx, cell in enumerate(row):
                header_key = (
                    active_header_keys[col_idx]
                    if col_idx < len(active_header_keys)
                    else f"col_{col_idx}"
                )

                cleaned_value = self._clean_cell_value(cell.value)
                if cleaned_value is None:
                    continue

                if header_key in position:
                    existing_value = position[header_key]
                    if isinstance(existing_value, list):
                        existing_value.append(cleaned_value)
                    else:
                        position[header_key] = [existing_value, cleaned_value]
                else:
                    position[header_key] = cleaned_value

                header_info = (
                    active_header_info[col_idx]
                    if col_idx < len(active_header_info)
                    else None
                )
                canonical_field = header_info["field"] if header_info else None

                if canonical_field and canonical_field not in canonical_row:
                    canonical_row[canonical_field] = cleaned_value

            summary_reason = self._detect_sum_row(position)
            if summary_reason:
                record_skip(row_idx, summary_reason)
                continue

            if not position:
                record_skip(row_idx, "no_data_after_cleaning")
                continue

            missing_fields = [
                field
                for field in required_fields
                if field not in canonical_row or canonical_row[field] in (None, "")
            ]
            if missing_fields:
                record_skip(
                    row_idx,
                    "missing_required_fields",
                    missing=missing_fields,
                    available=sorted(canonical_row.keys()),
                )
                continue

            if "unit" not in canonical_row:
                logger.debug(
                    "Row %s in sheet '%s' accepted without unit value",
                    row_idx,
                    sheet_name,
                )

            position['_source'] = f"sheet_{sheet_name}_row_{row_idx}"

            positions.append(position)
            last_data_row = row_idx

            logger.debug(
                "Accepted row %s in sheet '%s' with fields %s",
                row_idx,
                sheet_name,
                sorted(canonical_row.keys()),
            )

        logger.debug(f"Extracted {len(positions)} raw positions from sheet")

        sheet_diagnostics = {
            "header_found": bool(header_blocks),
            "header_row": first_header["row"] if first_header else None,
            "header_values": first_header["raw_headers"] if first_header else [],
            "matched_keywords": first_header["matched_fields"] if first_header else [],
            "searched_rows": sheet.max_row,
            "header_blocks": header_blocks,
            "skipped": skipped_entries,
            "summary": {
                "positions_found": len(positions),
                "positions_skipped": positions_skipped,
                "last_data_row": last_data_row,
                "rows_scanned": sheet.max_row,
            },
        }

        if not header_blocks:
            sheet_diagnostics["reason"] = "No header detected in sheet"
            sheet_diagnostics["best_candidate"] = best_candidate
            sheet_diagnostics["sample_rows"] = sample_rows

            logger.warning(
                "No header detected in sheet '%s'. Best candidate: %s",
                sheet_name,
                best_candidate,
            )

        return positions, sheet_diagnostics

    @staticmethod
    def _normalize_header_value(value: Any) -> str:
        """Normalize header values by lowering and stripping diacritics/symbols."""
        if value is None:
            return ""

        text = str(value).strip().lower()
        if not text:
            return ""

        normalized = unicodedata.normalize("NFKD", text)
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r"[^a-z0-9\s]", " ", normalized)
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized.strip()

    @staticmethod
    def _find_header_row(sheet: Worksheet) -> Dict[str, Any]:
        """
        Find the row that contains column headers

        Looks for rows with typical header keywords

        Returns:
            Detailed information about detected header row search
        """
        header_keywords = [
            "text",
            "popis",
            "opis",
            "nazev",
            "description",
            "desc",
            "jednotka",
            "mj",
            "m.j.",
            "množství",
            "mnozstvi",
            "qty",
            "quantity",
            "j.cena",
            "jednotkova cena",
            "jednotková cena",
            "cena celkem",
            "cena",
            "price",
            "kc",
            "czk",
            "kod",
            "kd",
            "typ",
            "normohodiny",
            "dph",
            "souhrn",
            "cena bez dph",
            "unit",
        ]

        normalized_keywords = {
            ExcelParser._normalize_header_value(keyword)
            for keyword in header_keywords
        }
        normalized_keywords.discard("")

        search_limit = min(100, sheet.max_row)
        sample_rows: List[List[str]] = []
        best_candidate: Optional[Dict[str, Any]] = None

        for row_idx in range(1, search_limit + 1):
            row = sheet[row_idx]
            raw_values = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in row
            ]

            if len(sample_rows) < 5:
                sample_rows.append(raw_values)

            normalized_cells = [
                ExcelParser._normalize_header_value(cell.value)
                for cell in row
                if cell.value is not None
            ]

            row_text = " ".join(value for value in normalized_cells if value)
            matched_keywords = {
                keyword for keyword in normalized_keywords if keyword in row_text
            }

            if matched_keywords:
                candidate = {
                    "row_index": row_idx,
                    "matches": sorted(matched_keywords),
                    "header_values": raw_values,
                }
                if not best_candidate or len(matched_keywords) > len(best_candidate["matches"]):
                    best_candidate = candidate

            if len(matched_keywords) >= 2:
                return {
                    "row_index": row_idx,
                    "header_values": raw_values,
                    "matched_keywords": sorted(matched_keywords),
                    "searched_rows": row_idx,
                    "sample_rows": sample_rows,
                    "best_candidate": best_candidate,
                }

        return {
            "row_index": None,
            "header_values": [],
            "matched_keywords": [],
            "searched_rows": search_limit,
            "sample_rows": sample_rows,
            "best_candidate": best_candidate,
        }
    
    @staticmethod
    def _clean_cell_value(value: Any) -> Optional[Any]:
        """Clean cell value by trimming strings and ignoring empty content."""
        if value is None:
            return None

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            return cleaned

        return value

    def _detect_header_candidate(self, normalized_values: List[str]) -> Optional[Dict[str, Any]]:
        """Detect whether a row is a header candidate based on recognized keywords."""
        matches: List[Tuple[int, str, str]] = []

        for idx, normalized in enumerate(normalized_values):
            if not normalized:
                continue

            canonical_field = self._map_header_to_field(normalized)
            if canonical_field:
                matches.append((idx, canonical_field, normalized))

        if not matches:
            return None

        matched_fields = sorted({field for _, field, _ in matches})

        return {
            "matches": [
                {
                    "index": idx,
                    "field": field,
                    "value": value,
                }
                for idx, field, value in matches
            ],
            "matched_fields": matched_fields,
            "match_count": len(matched_fields),
            "is_header": len(matched_fields) >= 3,
        }

    @staticmethod
    def _map_header_to_field(header: str) -> Optional[str]:
        """Map normalized header to canonical field name using keyword heuristics."""
        if not header:
            return None

        header_lower = header.lower()

        # Avoid treating unit price columns as units
        if "jednotkova" in header_lower and "cena" in header_lower:
            return None

        code_keywords = ["kod", "code", "oznaceni", "cislo", "c.", "signatura"]
        description_keywords = [
            "popis", "opis", "nazev", "description", "text", "polozka", "name",
            "prace", "item", "druh_prace",
        ]
        unit_keywords = ["jednotka", "mj", "m_j", "merna", "unit", "jedn."]
        quantity_keywords = [
            "mnozstvi", "quantity", "qty", "pocet", "mno", "mn", "qtty",
        ]

        normalized = header_lower.replace(" ", "_")

        for keyword in quantity_keywords:
            if keyword in normalized:
                return "quantity"

        for keyword in unit_keywords:
            if keyword in normalized:
                return "unit"

        for keyword in code_keywords:
            if keyword in normalized:
                return "code"

        for keyword in description_keywords:
            if keyword in normalized:
                return "description"

        return None

    @staticmethod
    def _detect_service_keyword(row_text: str) -> Optional[str]:
        """Detect if row contains known service keywords (summary, notes, etc.)."""
        if not row_text:
            return None

        keywords = {
            "rekapitulace": "rekapitulace",
            "souhrn": "souhrn",
            "soucet": "soucet",
            "součet": "součet",
            "sum": "sum",
            "subtotal": "subtotal",
            "summary": "summary",
            "celkem": "celkem",
            "spolu": "spolu",
            "pozn": "pozn",
            "poznám": "poznám",
            "note": "note",
        }

        for key, value in keywords.items():
            if key in row_text:
                return value

        return None

    @staticmethod
    def _detect_sum_row(position: Dict[str, Any]) -> Optional[str]:
        """Detect if a row represents a sum or total row based on its values."""
        if not position:
            return None

        all_values = " ".join(str(v) for v in position.values() if v)
        if not all_values:
            return None

        if re.match(r"^[-=*.\s]+$", all_values):
            return "separator_row"

        if all_values.isupper() and len(all_values) < 50:
            return "section_header"

        lowered = all_values.lower()
        if any(keyword in lowered for keyword in ["celkem", "sum", "souhrn", "rekapitulace", "spolu"]):
            return "sum_row"

        if len(all_values) < 3:
            return "short_row"

        return None
    
    def get_supported_extensions(self) -> set:
        """Return supported file extensions"""
        return {'.xlsx', '.xls'}


# ==============================================================================
# USAGE EXAMPLE
# ==============================================================================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        excel_path = Path(sys.argv[1])
        
        parser = ExcelParser()
        result = parser.parse(excel_path)
        
        print(f"\n📊 Parsing Results:")
        print(f"File: {result['document_info']['filename']}")
        print(f"Sheets: {result['document_info'].get('sheets', [])}")
        print(f"Positions found: {len(result['positions'])}")
        
        if result['positions']:
            print(f"\n📝 First 3 positions:")
            for pos in result['positions'][:3]:
                print(f"\n  Position: {pos.get('position_number', 'N/A')}")
                print(f"  Description: {pos.get('description', 'N/A')[:50]}")
                print(f"  Quantity: {pos.get('quantity', 0)} {pos.get('unit', '')}")
                if 'unit_price' in pos:
                    print(f"  Price: {pos['unit_price']} CZK")
    else:
        print("Usage: python excel_parser.py <path_to_excel>")
        print("Example: python excel_parser.py data/raw/project1/vykaz_vymer/estimate.xlsx")


registry.register_parser("excel", ExcelParser)
