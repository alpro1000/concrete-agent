"""
Memory-Efficient Parsers for 512MB RAM environment
ПОЛНАЯ ПОДДЕРЖКА ЧЕШСКИХ ФОРМАТОВ

Streaming parsing without loading entire files into memory
NO CLAUDE FALLBACK - only efficient local parsing
"""
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional, Iterator
import xml.etree.ElementTree as ET
import re

logger = logging.getLogger(__name__)

# ============================================================================
# MEMORY-EFFICIENT EXCEL PARSER - С ПОЛНОЙ ЧЕШСКОЙ ПОДДЕРЖКОЙ
# ============================================================================

class MemoryEfficientExcelParser:
    """
    Excel parser using streaming approach
    
    Memory usage: ~50-100MB regardless of file size
    Uses: openpyxl with read_only mode + iteration
    
    ✅ ПОЛНАЯ ПОДДЕРЖКА ЧЕШСКИХ ФОРМАТОВ:
    - PČ, Typ, Kód, Popis, MJ, Množství, J.cena, Cena celkem
    - Чешские числа: "1 220,168" → 1220.168
    - Case-insensitive поиск столбцов
    - Все варианты названий
    """
    
    # ✅ РАСШИРЕННЫЕ ЧЕШСКИЕ СТОЛБЦЫ - ВСЕ ВАРИАНТЫ!
    CZECH_COLUMN_KEYWORDS = {
        'number': [
            # Чешские варианты
            'pč', 'p.č.', 'p. č.', 'číslo', 'cislo', 'položka', 'polozka',
            # Английские
            'number', 'no', 'no.', '#', 'item', 'pos', 'position'
        ],
        'type': [
            'typ', 'type', 'druh', 'kategorie', 'category'
        ],
        'code': [
            # Чешские
            'kód', 'kod', 'kód položky', 'kod polozky', 
            # Английские
            'code', 'item code', 'position code'
        ],
        'description': [
            # Чешские
            'popis', 'název', 'nazev', 'popis práce', 'popis prace',
            'název práce', 'nazev prace', 'text',
            # Английские
            'description', 'name', 'title', 'work description'
        ],
        'unit': [
            # Чешские
            'mj', 'm.j.', 'm. j.', 'jednotka', 'měrná jednotka', 
            'merna jednotka', 'mj.',
            # Английские
            'unit', 'uom', 'measure', 'measurement unit'
        ],
        'quantity': [
            # Чешские
            'množství', 'mnozstvi', 'množství celkem', 'mnozstvi celkem',
            'počet', 'pocet', 'qty', 'objem',
            # Английские
            'quantity', 'amount', 'total quantity', 'volume'
        ],
        'unit_price': [
            # Чешские - ВСЕ ВАРИАНТЫ!
            'j.cena', 'j. cena', 'j . cena',  # С разными пробелами
            'J.cena', 'J. cena', 'J . cena',  # С заглавной
            'jednotková cena', 'jednotkova cena',
            'cena za mj', 'cena/mj', 'cena za jednotku',
            'cena jednotková', 'j cena',
            # Английские
            'unit price', 'price per unit', 'unit cost', 
            'price/unit', 'rate'
        ],
        'total_price': [
            # Чешские - ВСЕ ВАРИАНТЫ!
            'cena celkem', 'Cena celkem', 'CENA CELKEM',
            'celkem', 'Celkem', 'CELKEM',
            'celková cena', 'celkova cena', 'Celková cena',
            'cena celk.', 'cena celk', 'c. celkem',
            # Английские
            'total', 'total price', 'total cost', 
            'amount', 'total amount'
        ]
    }
    
    def __init__(self):
        """No Claude client - pure local parsing"""
        pass
    
    def parse(self, excel_path: Path, max_rows: int = 50000) -> Dict[str, Any]:
        """
        Parse Excel file using streaming
        
        Args:
            excel_path: Path to Excel file
            max_rows: Maximum rows to process (safety limit)
            
        Returns:
            Parsed data
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl required. Install: pip install openpyxl")
        
        logger.info(f"📊 Streaming Excel parse: {excel_path}")
        
        # ✅ Load in read_only mode (memory efficient!)
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        
        all_positions = []
        sections = []
        
        for sheet in wb.worksheets:
            logger.info(f"  Processing sheet: {sheet.title}")
            
            # Stream rows
            positions = self._parse_sheet_streaming(sheet, max_rows)
            
            if positions:
                all_positions.extend(positions)
                sections.append({
                    "name": sheet.title,
                    "positions_count": len(positions)
                })
        
        wb.close()  # ✅ Release memory!
        
        result = {
            "positions": all_positions,
            "total_positions": len(all_positions),
            "document_info": {
                "document_type": "Excel Estimate",
                "format": "STREAMING_EXCEL"
            },
            "sections": sections
        }
        
        logger.info(f"✅ Parsed {len(all_positions)} positions (streaming)")
        return result
    
    def _parse_sheet_streaming(self, sheet, max_rows: int) -> List[Dict[str, Any]]:
        """
        Parse sheet row by row (streaming)
        
        Memory: Only current row in memory!
        """
        positions = []
        header_row = None
        col_map = {}
        
        row_count = 0
        
        # Iterate through rows (streaming!)
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            row_count += 1
            
            # Safety limit
            if row_count > max_rows:
                logger.warning(f"⚠️ Reached max_rows limit ({max_rows})")
                break
            
            # Skip empty rows
            if not any(cell for cell in row):
                continue
            
            # Find header row
            if header_row is None:
                if self._is_header_row(row):
                    header_row = row
                    # ✅ CASE-INSENSITIVE + TRIM!
                    col_map = self._map_columns([
                        str(c).lower().strip() if c else '' 
                        for c in row
                    ])
                    logger.debug(f"  Header found at row {row_idx}")
                    logger.debug(f"  Column mapping: {col_map}")
                continue
            
            # Parse data row
            position = self._parse_row(row, col_map, row_idx)
            if position:
                positions.append(position)
        
        return positions
    
    def _is_header_row(self, row: tuple) -> bool:
        """Check if row is header"""
        row_text = ' '.join([str(c).lower() for c in row if c])
        
        # Must contain key column names
        czech_keywords = ['popis', 'kód', 'kod', 'mj', 'množství', 'mnozstvi']
        found = sum(1 for kw in czech_keywords if kw in row_text)
        
        return found >= 2
    
    def _map_columns(self, header: List[str]) -> Dict[str, int]:
        """
        Map column names to indices
        ✅ CASE-INSENSITIVE + TRIM + ВСЕ ВАРИАНТЫ
        """
        mapping = {}
        
        for field, keywords in self.CZECH_COLUMN_KEYWORDS.items():
            for idx, col in enumerate(header):
                # ✅ CASE-INSENSITIVE поиск!
                col_clean = col.lower().strip()
                
                for keyword in keywords:
                    keyword_clean = keyword.lower().strip()
                    
                    # Точное совпадение ИЛИ содержит
                    if keyword_clean == col_clean or keyword_clean in col_clean:
                        mapping[field] = idx
                        logger.debug(f"    Matched '{field}' → column {idx} ('{col}')")
                        break
                
                if field in mapping:
                    break
        
        return mapping
    
    def _parse_row(self, row: tuple, col_map: Dict[str, int], row_idx: int) -> Optional[Dict[str, Any]]:
        """Parse single row"""
        # Get description (required)
        desc_idx = col_map.get('description')
        if desc_idx is None or desc_idx >= len(row):
            return None
        
        description = row[desc_idx]
        if not description:
            return None
        
        # Skip header-like descriptions
        desc_lower = str(description).lower().strip()
        if any(kw in desc_lower for kw in ['popis', 'description', 'název', 'name']):
            if len(desc_lower) < 15:  # Короткое = вероятно header
                return None
        
        def get_cell(field: str, default=''):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return default
            return row[idx] if row[idx] is not None else default
        
        position = {
            "number": self._parse_int(get_cell('number', '')),
            "type": str(get_cell('type', '')).strip(),
            "code": str(get_cell('code', '')).strip(),
            "description": str(description).strip(),
            "unit": str(get_cell('unit', '')).strip(),
            "quantity": self._parse_czech_number(get_cell('quantity', '0')),
            "unit_price": self._parse_czech_number(get_cell('unit_price', '0')),
            "total_price": self._parse_czech_number(get_cell('total_price', '0')),
            "row_number": row_idx + 1
        }
        
        # Calculate total if missing
        if position["quantity"] and position["unit_price"] and not position["total_price"]:
            position["total_price"] = position["quantity"] * position["unit_price"]
        
        return position
    
    def _parse_czech_number(self, value) -> float:
        """
        Parse Czech number format
        ✅ УЛУЧШЕННЫЙ: обрабатывает ВСЕ варианты
        
        Examples:
        - "1 220,168" → 1220.168
        - "1.220,168" → 1220.168 (европейский формат)
        - "1220.168" → 1220.168 (английский формат)
        - "15,50" → 15.50
        - "1 000" → 1000.0
        """
        if isinstance(value, (int, float)):
            return float(value)
        
        if not value:
            return 0.0
        
        try:
            text = str(value).strip()
            
            # Удалить пробелы (разделители тысяч)
            text = text.replace(' ', '').replace('\xa0', '')
            
            # Удалить валюту
            text = text.replace('Kč', '').replace('CZK', '').replace('EUR', '').replace('€', '')
            text = text.strip()
            
            # Определить формат числа
            # Если есть И точка И запятая → европейский формат
            if '.' in text and ',' in text:
                # Европейский: 1.220,168 → запятая = десятичная
                text = text.replace('.', '')  # Удалить точки (тысячи)
                text = text.replace(',', '.')  # Запятая → точка (десятичная)
            elif ',' in text:
                # Только запятая → может быть десятичная (чешский формат)
                # Проверить: если после запятой 3+ цифры → это тысячи, иначе десятичная
                parts = text.split(',')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    # Десятичная: 15,50
                    text = text.replace(',', '.')
                else:
                    # Тысячи: 1,220 → просто удалить
                    text = text.replace(',', '')
            # Если только точка → может быть десятичная (английский) или тысячи
            # Оставляем как есть
            
            # Удалить все кроме цифр, точки и минуса
            text = re.sub(r'[^\d.-]', '', text)
            
            if not text or text == '-':
                return 0.0
            
            return float(text)
        except (ValueError, AttributeError) as e:
            logger.debug(f"Failed to parse number '{value}': {e}")
            return 0.0
    
    def _parse_int(self, value) -> int:
        """Parse integer"""
        try:
            return int(self._parse_czech_number(value))
        except:
            return 0


# ============================================================================
# MEMORY-EFFICIENT XML PARSER - С ЧЕШСКОЙ ПОДДЕРЖКОЙ
# ============================================================================

class MemoryEfficientXMLParser:
    """
    XML parser using iterparse (streaming)
    
    Memory usage: ~20-50MB regardless of file size
    Uses: ET.iterparse for streaming XML
    """
    
    def __init__(self):
        """No Claude client - pure local parsing"""
        pass
    
    def parse(self, xml_path: Path) -> Dict[str, Any]:
        """
        Parse XML using streaming iterparse
        
        Args:
            xml_path: Path to XML file
            
        Returns:
            Parsed data
        """
        logger.info(f"📄 Streaming XML parse: {xml_path}")
        
        # Detect format first (quick scan)
        format_type = self._detect_format_quick(xml_path)
        logger.info(f"  Format: {format_type}")
        
        if format_type == "KROS_TABLE":
            return self._parse_table_streaming(xml_path)
        else:
            return self._parse_generic_streaming(xml_path)
    
    def _detect_format_quick(self, xml_path: Path) -> str:
        """Quick format detection (first 1KB)"""
        with open(xml_path, 'r', encoding='utf-8') as f:
            sample = f.read(1024).lower()
        
        if '<tz>' in sample or '<row>' in sample:
            return "KROS_TABLE"
        return "GENERIC"
    
    def _parse_table_streaming(self, xml_path: Path) -> Dict[str, Any]:
        """
        Parse KROS Table XML using streaming
        
        Memory: Only current element in memory!
        """
        positions = []
        
        # ✅ Streaming parse!
        for event, elem in ET.iterparse(xml_path, events=('end',)):
            if elem.tag == 'Row':
                position = self._parse_table_row_elem(elem)
                if position:
                    positions.append(position)
                
                # ✅ Clear element to free memory!
                elem.clear()
        
        result = {
            "positions": positions,
            "total_positions": len(positions),
            "document_info": {
                "document_type": "KROS Table XML",
                "format": "STREAMING_XML"
            },
            "sections": []
        }
        
        logger.info(f"✅ Parsed {len(positions)} positions (streaming)")
        return result
    
    def _parse_table_row_elem(self, elem: ET.Element) -> Optional[Dict[str, Any]]:
        """Parse single Row element"""
        try:
            position = {
                "code": self._get_text(elem, 'C1'),
                "description": self._get_text(elem, 'C2'),
                "unit": self._get_text(elem, 'C3'),
                "quantity": self._parse_float(self._get_text(elem, 'C4')),
                "unit_price": self._parse_float(self._get_text(elem, 'C5')),
                "total_price": self._parse_float(self._get_text(elem, 'C6'))
            }
            
            if position["description"] or position["code"]:
                return position
            
            return None
        except Exception as e:
            logger.debug(f"Failed to parse row: {e}")
            return None
    
    def _parse_generic_streaming(self, xml_path: Path) -> Dict[str, Any]:
        """Parse generic XML using streaming"""
        positions = []
        current_elem = None
        
        for event, elem in ET.iterparse(xml_path, events=('start', 'end')):
            if event == 'start':
                if self._looks_like_position_tag(elem.tag):
                    current_elem = elem
            
            elif event == 'end' and current_elem is not None and elem == current_elem:
                position = self._extract_position_from_elem(elem)
                if position:
                    positions.append(position)
                
                elem.clear()
                current_elem = None
        
        result = {
            "positions": positions,
            "total_positions": len(positions),
            "document_info": {
                "document_type": "Generic XML",
                "format": "STREAMING_XML"
            },
            "sections": []
        }
        
        logger.info(f"✅ Parsed {len(positions)} positions (streaming)")
        return result
    
    def _looks_like_position_tag(self, tag: str) -> bool:
        """Check if tag looks like a position"""
        tag_lower = tag.lower()
        keywords = ['item', 'position', 'row', 'polozka', 'pozice']
        return any(kw in tag_lower for kw in keywords)
    
    def _extract_position_from_elem(self, elem: ET.Element) -> Optional[Dict[str, Any]]:
        """Extract position data from element"""
        position = {}
        
        for child in elem:
            tag = child.tag.lower()
            text = child.text or ''
            
            if 'code' in tag or 'kod' in tag:
                position['code'] = text
            elif 'popis' in tag or 'description' in tag or 'name' in tag:
                position['description'] = text
            elif 'mj' in tag or 'unit' in tag:
                position['unit'] = text
            elif 'mnozstvi' in tag or 'quantity' in tag:
                position['quantity'] = self._parse_float(text)
            elif 'cena' in tag and 'jednotkova' in tag:
                position['unit_price'] = self._parse_float(text)
        
        if position.get('description') or position.get('code'):
            return position
        
        return None
    
    def _get_text(self, elem: ET.Element, tag: str) -> str:
        """Get text from child element"""
        child = elem.find(tag)
        return child.text if child is not None and child.text else ''
    
    def _parse_float(self, text: str) -> float:
        """Parse float from text"""
        try:
            text = text.strip().replace(' ', '').replace(',', '.')
            return float(text) if text else 0.0
        except:
            return 0.0


# ============================================================================
# MEMORY-EFFICIENT PDF PARSER
# ============================================================================

class MemoryEfficientPDFParser:
    """
    PDF parser processing page by page
    
    Memory usage: ~100MB regardless of PDF size
    Uses: pdfplumber with page-by-page processing
    """
    
    def __init__(self):
        """No Claude client - pure local parsing"""
        pass
    
    def parse(self, pdf_path: Path, max_pages: int = 100) -> Dict[str, Any]:
        """
        Parse PDF page by page (streaming)
        
        Args:
            pdf_path: Path to PDF file
            max_pages: Maximum pages to process
            
        Returns:
            Parsed data
        """
        try:
            import pdfplumber
        except ImportError:
            raise ImportError("pdfplumber required. Install: pip install pdfplumber")
        
        logger.info(f"📕 Streaming PDF parse: {pdf_path}")
        
        positions = []
        
        # ✅ Open PDF (but don't load all at once)
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            pages_to_process = min(total_pages, max_pages)
            
            logger.info(f"  Processing {pages_to_process} of {total_pages} pages")
            
            # Process each page individually
            for page_num in range(pages_to_process):
                # ✅ Load only ONE page at a time!
                page = pdf.pages[page_num]
                
                # Extract tables from this page
                tables = page.extract_tables()
                
                for table_num, table in enumerate(tables):
                    parsed = self._parse_table_memory_safe(table, page_num + 1, table_num + 1)
                    positions.extend(parsed)
                
                # ✅ Explicitly release page (hint to GC)
                del page
        
        result = {
            "positions": positions,
            "total_positions": len(positions),
            "document_info": {
                "document_type": "PDF Estimate",
                "format": "STREAMING_PDF",
                "pages_processed": pages_to_process
            },
            "sections": []
        }
        
        logger.info(f"✅ Parsed {len(positions)} positions from {pages_to_process} pages")
        return result
    
    def _parse_table_memory_safe(self, table: List[List[str]], page: int, table_num: int) -> List[Dict[str, Any]]:
        """Parse table without loading extra data"""
        if not table or len(table) < 2:
            return []
        
        positions = []
        
        # Header row
        header = [str(c).lower().strip() if c else '' for c in table[0]]
        
        # Find columns - ✅ УЛУЧШЕННЫЙ ПОИСК!
        col_map = {}
        keywords = {
            'code': ['kód', 'kod', 'code', 'položka'],
            'description': ['popis', 'description', 'název', 'nazev'],
            'unit': ['mj', 'jednotka', 'unit', 'm.j.'],
            'quantity': ['množství', 'mnozstvi', 'quantity', 'počet'],
            'unit_price': ['j.cena', 'j. cena', 'jednotková', 'cena', 'unit price']
        }
        
        for field, kws in keywords.items():
            for idx, col in enumerate(header):
                if any(kw in col for kw in kws):
                    col_map[field] = idx
                    break
        
        # Parse rows
        for row in table[1:]:
            if not row:
                continue
            
            desc_idx = col_map.get('description')
            if desc_idx is None or desc_idx >= len(row):
                continue
            
            description = row[desc_idx]
            if not description:
                continue
            
            def get(field, default=''):
                idx = col_map.get(field)
                if idx is None or idx >= len(row):
                    return default
                return row[idx] if row[idx] else default
            
            position = {
                "code": str(get('code')),
                "description": str(description),
                "unit": str(get('unit')),
                "quantity": self._parse_float(get('quantity', '0')),
                "unit_price": self._parse_float(get('unit_price', '0')),
                "page": page,
                "table": table_num
            }
            
            positions.append(position)
        
        return positions
    
    def _parse_float(self, text: str) -> float:
        """Parse float"""
        try:
            text = str(text).strip().replace(' ', '').replace(',', '.')
            text = text.replace('Kč', '').replace('CZK', '')
            return float(text) if text else 0.0
        except:
            return 0.0
