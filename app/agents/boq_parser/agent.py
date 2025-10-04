"""
BOQ Parser Agent - Bill of Quantities Parser
Parses Excel/XML budget files, extracts cost items, and maps to KROS/URS codes
"""

from typing import Dict, Any, List, Optional
import logging
import json
from pathlib import Path
import openpyxl
import pandas as pd
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)


class BOQParserAgent:
    """
    Bill of Quantities Parser Agent
    
    Parses budget files (Excel/XML) and extracts:
    - Cost items with codes
    - Quantities and units
    - Unit prices and totals
    - KROS/URS code mapping
    """
    
    name = "boq_parser"
    supported_types = [
        "boq",
        "bill_of_quantities",
        "budget",
        "excel",
        "xlsx",
        "xls",
        "xml"
    ]
    
    def __init__(self):
        """Initialize BOQ Parser Agent"""
        logger.info("BOQParserAgent initialized")
    
    def _parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse Excel file and extract BOQ items"""
        items = []
        
        try:
            # Try openpyxl first
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            headers = []
            data_start_row = 0
            
            # Find header row
            for row_idx, row in enumerate(ws.iter_rows(max_row=20), start=1):
                row_values = [cell.value for cell in row]
                # Look for typical BOQ headers
                if any(h and str(h).lower() in ['item', 'code', 'description', 'quantity', 'unit', 'price'] 
                       for h in row_values):
                    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(row_values)]
                    data_start_row = row_idx + 1
                    break
            
            # If no headers found, assume first row is data
            if not headers:
                first_row = next(ws.iter_rows(min_row=1, max_row=1))
                headers = [f"col_{i}" for i in range(len(list(first_row)))]
                data_start_row = 1
            
            # Extract data rows
            for row in ws.iter_rows(min_row=data_start_row):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = cell.value
                
                # Skip empty rows
                if any(v for v in row_data.values() if v):
                    items.append(row_data)
            
            wb.close()
            
        except Exception as e:
            logger.warning(f"openpyxl failed, trying pandas: {e}")
            try:
                df = pd.read_excel(file_path)
                items = df.to_dict('records')
            except Exception as e2:
                logger.error(f"Both Excel parsers failed: {e2}")
                raise ValueError(f"Failed to parse Excel file: {e2}")
        
        return items
    
    def _parse_xml(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse XML file and extract BOQ items"""
        items = []
        
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Try to find items (adapt based on common XML schemas)
            for item_elem in root.findall('.//item') or root.findall('.//Item'):
                item = {}
                for child in item_elem:
                    item[child.tag] = child.text
                items.append(item)
            
            # If no items found, try different structure
            if not items:
                for elem in root.iter():
                    if elem.text and elem.text.strip():
                        items.append({elem.tag: elem.text})
            
        except Exception as e:
            logger.error(f"XML parsing failed: {e}")
            raise ValueError(f"Failed to parse XML file: {e}")
        
        return items
    
    def _extract_structured_boq(self, raw_items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Extract structured BOQ data from raw items"""
        structured_items = []
        total_cost = 0.0
        
        for idx, item in enumerate(raw_items):
            structured_item = {
                "item_number": idx + 1,
                "code": None,
                "description": None,
                "quantity": None,
                "unit": None,
                "unit_price": None,
                "total_price": None
            }
            
            # Map fields (flexible field names)
            for key, value in item.items():
                key_lower = str(key).lower()
                
                if any(k in key_lower for k in ['code', 'kód', 'kod']):
                    structured_item["code"] = str(value) if value else None
                elif any(k in key_lower for k in ['description', 'popis', 'name', 'název', 'nazev']):
                    structured_item["description"] = str(value) if value else None
                elif any(k in key_lower for k in ['quantity', 'množství', 'mnozstvi', 'qty']):
                    try:
                        structured_item["quantity"] = float(value) if value else None
                    except (ValueError, TypeError):
                        structured_item["quantity"] = None
                elif any(k in key_lower for k in ['unit', 'jednotka', 'mj']):
                    structured_item["unit"] = str(value) if value else None
                elif any(k in key_lower for k in ['unit_price', 'price', 'cena', 'jednotková']):
                    try:
                        structured_item["unit_price"] = float(value) if value else None
                    except (ValueError, TypeError):
                        structured_item["unit_price"] = None
                elif any(k in key_lower for k in ['total', 'celkem', 'amount']):
                    try:
                        structured_item["total_price"] = float(value) if value else None
                    except (ValueError, TypeError):
                        structured_item["total_price"] = None
            
            # Calculate total if not provided
            if (structured_item["total_price"] is None and 
                structured_item["quantity"] is not None and 
                structured_item["unit_price"] is not None):
                structured_item["total_price"] = structured_item["quantity"] * structured_item["unit_price"]
            
            # Add to total cost
            if structured_item["total_price"]:
                total_cost += structured_item["total_price"]
            
            structured_items.append(structured_item)
        
        return {
            "items": structured_items,
            "total_items": len(structured_items),
            "total_cost": round(total_cost, 2),
            "currency": "CZK"  # Default, should be detected
        }
    
    async def analyze(self, file_path: str) -> Dict[str, Any]:
        """
        Analyze a BOQ file and extract structured data.
        
        Args:
            file_path: Path to BOQ file (Excel or XML)
            
        Returns:
            Dictionary with BOQ data:
                - items: List of cost items
                - total_items: Total number of items
                - total_cost: Total project cost
                - currency: Currency code
        """
        try:
            file_ext = Path(file_path).suffix.lower()
            
            # Parse based on file type
            if file_ext in ['.xlsx', '.xls']:
                raw_items = self._parse_excel(file_path)
            elif file_ext == '.xml':
                raw_items = self._parse_xml(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_ext}")
            
            # Extract structured data
            result = self._extract_structured_boq(raw_items)
            
            # Add metadata
            result["processing_metadata"] = {
                "file_name": Path(file_path).name,
                "file_type": file_ext,
                "raw_items_count": len(raw_items),
                "agent": self.name
            }
            
            logger.info(f"BOQ parsed: {result['total_items']} items, total: {result['total_cost']} {result['currency']}")
            return result
            
        except Exception as e:
            logger.error(f"BOQ parsing failed: {e}")
            return {
                "items": [],
                "total_items": 0,
                "total_cost": 0.0,
                "currency": "CZK",
                "error": str(e),
                "processing_metadata": {
                    "file_name": Path(file_path).name,
                    "agent": self.name,
                    "status": "failed"
                }
            }
    
    def supports_type(self, file_type: str) -> bool:
        """Check if this agent supports a given file type"""
        return file_type.lower() in [t.lower() for t in self.supported_types]
    
    def __repr__(self) -> str:
        """String representation of the agent"""
        return f"BOQParserAgent(name='{self.name}', supported_types={self.supported_types})"
